"""Excel export / import / template generation for class groups."""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from ..models import User, Student, Project, PhaseProgress, PHASE_ORDER


# ── Colour constants ───────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill('solid', fgColor='1F3864')
_STUDENT_FILL = PatternFill('solid', fgColor='D9E1F2')
_GREEN_FILL   = PatternFill('solid', fgColor='C6EFCE')
_RED_FILL     = PatternFill('solid', fgColor='FFC7CE')
_GRAY_FILL    = PatternFill('solid', fgColor='D9D9D9')

_HEADER_FONT  = Font(color='FFFFFF', bold=True, name='Calibri')
_BOLD_FONT    = Font(bold=True, name='Calibri')
_CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
_RIGHT        = Alignment(horizontal='right',  vertical='center', wrap_text=True)

_PHASE_HEBREW = {
    'initiation':       'פתיחה',
    'characterization': 'אפיון',
    'analysis_design':  'ניתוח ועיצוב',
    'implementation':   'מימוש',
    'testing':          'בדיקות',
    'final_submission': 'הגשה סופית',
}


# ── Export ─────────────────────────────────────────────────────────────────────

def export_class_to_excel(class_group, include_comments=True) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_group.name} {class_group.year}"

    # Build headers
    headers = ['שם תלמיד', 'שם פרויקט', 'נושא']
    for phase in PHASE_ORDER:
        headers.append(f"{_PHASE_HEBREW[phase]} %")
        if include_comments:
            headers.append(f"הערות {_PHASE_HEBREW[phase]}")
    headers.append('התקדמות כללית %')

    # Write + style header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER

    # Write student rows
    for row_idx, student in enumerate(class_group.students, 2):
        col = 1

        # Student name (col A)
        name_cell = ws.cell(row=row_idx, column=col, value=student.user.full_name)
        name_cell.fill      = _STUDENT_FILL
        name_cell.font      = _BOLD_FONT
        name_cell.alignment = _RIGHT
        col += 1

        if student.project:
            ws.cell(row=row_idx, column=col, value=student.project.title).alignment = _RIGHT
            col += 1
            ws.cell(row=row_idx, column=col, value=student.project.subject).alignment = _RIGHT
            col += 1

            phases_dict = {pp.phase: pp for pp in student.project.phases.all()}
            percentages = []
            for phase in PHASE_ORDER:
                pp  = phases_dict.get(phase)
                pct = pp.percentage if pp else 0
                percentages.append(pct)

                pct_cell = ws.cell(row=row_idx, column=col, value=pct)
                pct_cell.alignment = _CENTER
                if pct == 100:
                    pct_cell.fill = _GREEN_FILL
                elif pct == 0:
                    pct_cell.fill = _RED_FILL
                col += 1

                if include_comments:
                    notes_cell = ws.cell(row=row_idx, column=col,
                                         value=(pp.notes if pp and pp.notes else ''))
                    notes_cell.alignment = _RIGHT
                    col += 1

            overall = sum(percentages) // len(percentages) if percentages else 0
            overall_cell = ws.cell(row=row_idx, column=col, value=overall)
            overall_cell.alignment = _CENTER
            overall_cell.font = _BOLD_FONT
            if overall == 100:
                overall_cell.fill = _GREEN_FILL
        else:
            col += 2  # project title + subject blank
            for phase in PHASE_ORDER:
                ws.cell(row=row_idx, column=col)
                col += 1
                if include_comments:
                    ws.cell(row=row_idx, column=col)
                    col += 1
            ws.cell(row=row_idx, column=col)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    for i in range(4, len(headers) + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 12
    # Last column (overall)
    ws.column_dimensions[get_column_letter(len(headers))].width = 14

    ws.freeze_panes = 'B2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Import template ────────────────────────────────────────────────────────────

def generate_import_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'תבנית ייבוא'

    headers = ['שם מלא', 'שם משתמש', 'אימייל', 'סיסמה', 'כיתה', 'שם פרויקט', 'נושא פרויקט']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER

    sample_rows = [
        ['ישראל ישראלי', 'israel123', 'israel@school.local', 'pass123', "י'1", 'מערכת ניטור', 'אבטחת מידע'],
        ['שרה כהן',      'sarah456',  'sarah@school.local',  'pass456', "י'1", 'אפליקציית הגנה', 'סייבר'],
        ['דוד לוי',       'david789',  'david@school.local',  'pass789', "יא'2", '', ''],
    ]
    for row_data in sample_rows:
        row_idx = sample_rows.index(row_data) + 2
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = _GRAY_FILL
            cell.alignment = _RIGHT

    col_widths = [25, 18, 30, 15, 10, 25, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Import ─────────────────────────────────────────────────────────────────────

def import_students_from_excel(file_stream, teacher_user):
    """
    Returns (imported: list[dict], errors: list[dict]).
    Each error dict has {row, username, reason}.
    """
    from openpyxl import load_workbook
    from ..extensions import db
    from ..models import ClassGroup

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    imported = []
    errors   = []
    current_year = datetime.utcnow().year

    # Cache class groups created this session
    class_cache = {}

    for row_num, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            continue

        full_name    = str(row[0]).strip() if row[0] else ''
        username     = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        email        = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        password     = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        class_name   = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        project_title = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        project_subj  = str(row[6]).strip() if len(row) > 6 and row[6] else ''

        if not username or not full_name:
            errors.append({'row': row_num, 'username': username or '(ריק)', 'reason': 'שם משתמש או שם מלא חסרים'})
            continue

        if not email:
            email = f"{username}@school.local"

        if not password:
            password = username + '123'

        if User.query.filter_by(username=username).first():
            errors.append({'row': row_num, 'username': username, 'reason': 'שם משתמש קיים כבר'})
            continue

        if User.query.filter_by(email=email).first():
            errors.append({'row': row_num, 'username': username, 'reason': 'אימייל קיים כבר'})
            continue

        # Resolve or create class group
        class_group = None
        if class_name:
            cache_key = (class_name, current_year)
            if cache_key in class_cache:
                class_group = class_cache[cache_key]
            else:
                class_group = ClassGroup.query.filter_by(
                    name=class_name, year=current_year, teacher_id=teacher_user.id
                ).first()
                if not class_group:
                    class_group = ClassGroup(
                        name=class_name, year=current_year, teacher_id=teacher_user.id
                    )
                    db.session.add(class_group)
                    db.session.flush()
                class_cache[cache_key] = class_group

        user = User(username=username, full_name=full_name, email=email, role='student')
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        student = Student(
            user_id=user.id,
            class_name=class_name,
            class_group_id=class_group.id if class_group else None,
            teacher_id=teacher_user.id,
        )
        db.session.add(student)
        db.session.flush()

        if project_title:
            subj = project_subj or 'כללי'
            project = Project(student_id=student.id, title=project_title, subject=subj)
            db.session.add(project)
            db.session.flush()
            for phase in PHASE_ORDER:
                db.session.add(PhaseProgress(
                    project_id=project.id, phase=phase, percentage=0,
                    updated_by_id=teacher_user.id,
                ))

        db.session.commit()
        imported.append({'row': row_num, 'username': username, 'full_name': full_name})

    return imported, errors


# ── Phase-status template (pre-filled) ────────────────────────────────────────

def generate_phase_template(teacher_user) -> BytesIO:
    """
    Download a pre-filled template with every student's current phase %.
    Teacher edits the numbers and re-uploads.
    Columns: שם משתמש | שם תלמיד | כיתה | שם פרויקט | [phase %] x6 | [הערות] x6
    Column A (username) is the key used during import.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'סטטוס שלבים'

    # Build header list
    headers = ['שם משתמש', 'שם תלמיד', 'כיתה', 'שם פרויקט']
    for phase in PHASE_ORDER:
        headers.append(f"{_PHASE_HEBREW[phase]} %")
        headers.append(f"הערות {_PHASE_HEBREW[phase]}")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER

    # Fetch all students who belong to this teacher's students
    # (all students in the system — teacher sees everyone)
    students = (Student.query
                .join(User)
                .order_by(User.full_name)
                .all())

    for row_idx, student in enumerate(students, 2):
        phases_dict = {}
        if student.project:
            phases_dict = {pp.phase: pp for pp in student.project.phases.all()}

        class_label = (student.class_group.display_name
                       if student.class_group
                       else (student.class_name or ''))

        col = 1
        # A: username (key)
        c = ws.cell(row=row_idx, column=col, value=student.user.username)
        c.font = _BOLD_FONT; c.alignment = _RIGHT
        col += 1
        # B: full name (read-only hint — gray)
        c = ws.cell(row=row_idx, column=col, value=student.user.full_name)
        c.fill = _GRAY_FILL; c.alignment = _RIGHT
        col += 1
        # C: class
        c = ws.cell(row=row_idx, column=col, value=class_label)
        c.fill = _GRAY_FILL; c.alignment = _RIGHT
        col += 1
        # D: project title
        c = ws.cell(row=row_idx, column=col,
                    value=student.project.title if student.project else '— אין פרויקט —')
        c.fill = _GRAY_FILL; c.alignment = _RIGHT
        col += 1

        for phase in PHASE_ORDER:
            pp  = phases_dict.get(phase)
            pct = pp.percentage if pp else 0
            notes = pp.notes if pp and pp.notes else ''

            pct_cell = ws.cell(row=row_idx, column=col, value=pct)
            pct_cell.alignment = _CENTER
            if pct == 100:
                pct_cell.fill = _GREEN_FILL
            elif pct == 0 and student.project:
                pct_cell.fill = _RED_FILL
            col += 1

            notes_cell = ws.cell(row=row_idx, column=col, value=notes)
            notes_cell.alignment = _RIGHT
            col += 1

    # Column widths
    ws.column_dimensions['A'].width = 18  # username
    ws.column_dimensions['B'].width = 22  # full name
    ws.column_dimensions['C'].width = 12  # class
    ws.column_dimensions['D'].width = 28  # project
    for i in range(5, len(headers) + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 12

    ws.freeze_panes = 'E2'  # freeze name/class/project columns

    # Add a note row at the top explaining read-only columns
    ws.sheet_properties.tabColor = '4472C4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Phase-status import ────────────────────────────────────────────────────────

def import_phases_from_excel(file_stream, teacher_user):
    """
    Update PhaseProgress rows from an uploaded Excel file.
    Identifies students by username (col A).
    Returns (updated: list[dict], errors: list[dict]).
    """
    from openpyxl import load_workbook
    from ..extensions import db

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    updated = []
    errors  = []

    for row_num, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            continue

        username = str(row[0]).strip() if row[0] else ''
        if not username:
            errors.append({'row': row_num, 'username': '(ריק)', 'reason': 'שם משתמש חסר'})
            continue

        user = User.query.filter_by(username=username).first()
        if not user or not user.student_profile:
            errors.append({'row': row_num, 'username': username, 'reason': 'תלמיד לא נמצא'})
            continue

        student = user.student_profile
        if not student.project:
            errors.append({'row': row_num, 'username': username, 'reason': 'אין פרויקט לתלמיד'})
            continue

        phases_dict = {pp.phase: pp for pp in student.project.phases.all()}
        changes = []

        # Phase data starts at col index 4 (0-based), pairs of (%, notes)
        col_offset = 4
        for phase_idx, phase in enumerate(PHASE_ORDER):
            pct_col   = col_offset + phase_idx * 2
            notes_col = pct_col + 1

            raw_pct = row[pct_col] if len(row) > pct_col else None
            raw_notes = row[notes_col] if len(row) > notes_col else None

            if raw_pct is None:
                continue  # skip blank — don't overwrite

            try:
                pct = max(0, min(100, int(float(str(raw_pct)))))
            except (ValueError, TypeError):
                errors.append({'row': row_num, 'username': username,
                               'reason': f"ערך לא תקין בעמודת {_PHASE_HEBREW[phase]}: {raw_pct}"})
                continue

            notes = str(raw_notes).strip() if raw_notes is not None else None

            pp = phases_dict.get(phase)
            if pp:
                pp.percentage    = pct
                if notes is not None:
                    pp.notes     = notes
                pp.updated_by_id = teacher_user.id
                pp.updated_at    = datetime.utcnow()
            else:
                # Create missing PhaseProgress row
                pp = PhaseProgress(
                    project_id=student.project.id,
                    phase=phase,
                    percentage=pct,
                    notes=notes or '',
                    updated_by_id=teacher_user.id,
                )
                db.session.add(pp)

            changes.append(f"{_PHASE_HEBREW[phase]}: {pct}%")

        student.project.updated_at = datetime.utcnow()
        db.session.commit()
        updated.append({
            'row':       row_num,
            'username':  username,
            'full_name': user.full_name,
            'changes':   ', '.join(changes) if changes else 'ללא שינוי',
        })

    return updated, errors
